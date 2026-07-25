# dashboard/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Sum, Avg, Count
from datetime import datetime, date
import csv
from django.http import HttpResponse
from decimal import Decimal
from products.models import Product, Category
from orders.models import Order, OrderItem
from products.forms import ProductForm, CategoryForm

# ============ FONCTION ADMIN (DÉFINIE EN PREMIER) ============
def is_admin(user):
    """Vérifie si l'utilisateur est un administrateur"""
    return user.is_superuser or user.is_staff

# ============ DASHBOARD PRINCIPAL ============
@login_required
@user_passes_test(is_admin)
def index(request):
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    total_orders = Order.objects.count()
    total_products = Product.objects.count()
    total_users = User.objects.count()
    total_categories = Category.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    
    total_revenue = Order.objects.filter(status='confirmed').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    average_order = Order.objects.filter(status='confirmed').aggregate(Avg('total_amount'))['total_amount__avg'] or 0
    
    orders_this_month = Order.objects.filter(created_at__date__gte=first_day_of_month).count()
    revenue_this_month = Order.objects.filter(
        created_at__date__gte=first_day_of_month,
        status='confirmed'
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    top_products = OrderItem.objects.values('product__name').annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')[:5]
    
    context = {
        'total_orders': total_orders,
        'total_products': total_products,
        'total_users': total_users,
        'total_categories': total_categories,
        'pending_orders': pending_orders,
        'total_revenue': total_revenue,
        'average_order': average_order,
        'orders_this_month': orders_this_month,
        'revenue_this_month': revenue_this_month,
        'top_products': top_products,
        'recent_orders': Order.objects.order_by('-created_at')[:10],
    }
    return render(request, 'dashboard/index.html', context)

# ============ GESTION DES PRODUITS ============
@login_required
@user_passes_test(is_admin)
def products(request):
    products = Product.objects.all()
    return render(request, 'dashboard/products.html', {'products': products})

@login_required
@user_passes_test(is_admin)
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Produit créé avec succès !')
            return redirect('dashboard:products')
    else:
        form = ProductForm()
    return render(request, 'dashboard/product_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Produit modifié avec succès !')
            return redirect('dashboard:products')
    else:
        form = ProductForm(instance=product)
    return render(request, 'dashboard/product_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, '✅ Produit supprimé avec succès !')
        return redirect('dashboard:products')
    return render(request, 'dashboard/product_confirm_delete.html', {'product': product})

# ============ GESTION DES COMMANDES ============
@login_required
@user_passes_test(is_admin)
def orders(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'dashboard/orders.html', {'orders': orders})

@login_required
@user_passes_test(is_admin)
def order_status(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES):
            order.status = new_status
            order.save()
            messages.success(request, f'✅ Statut de la commande #{order.order_number} mis à jour !')
    return redirect('dashboard:orders')

# ============ GESTION DES UTILISATEURS ============
@login_required
@user_passes_test(is_admin)
def users(request):
    users = User.objects.all()
    return render(request, 'dashboard/users.html', {'users': users})

@login_required
@user_passes_test(is_admin)
def user_toggle_status(request, pk):
    user = get_object_or_404(User, pk=pk)
    if not user.is_superuser:
        user.is_active = not user.is_active
        user.save()
        status = "activé" if user.is_active else "désactivé"
        messages.success(request, f'✅ Utilisateur {user.username} {status} avec succès !')
    return redirect('dashboard:users')

# ============ GESTION DES CATÉGORIES ============
@login_required
@user_passes_test(is_admin)
def categories(request):
    categories = Category.objects.all()
    return render(request, 'dashboard/categories.html', {'categories': categories})

@login_required
@user_passes_test(is_admin)
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Catégorie créée avec succès !')
            return redirect('dashboard:categories')
    else:
        form = CategoryForm()
    return render(request, 'dashboard/category_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Catégorie modifiée avec succès !')
            return redirect('dashboard:categories')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'dashboard/category_form.html', {'form': form})

@login_required
@user_passes_test(is_admin)
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, '✅ Catégorie supprimée avec succès !')
        return redirect('dashboard:categories')
    return render(request, 'dashboard/category_confirm_delete.html', {'category': category})

# ============ EXPORT CSV ============
@login_required
@user_passes_test(is_admin)
def export_orders_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="commandes_lozashop.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Numéro', 'Client', 'Email', 'Total', 'Statut', 
        'Date', 'Adresse de livraison', 'Paiement'
    ])
    
    orders = Order.objects.all().order_by('-created_at')
    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.username,
            order.user.email or '',
            float(order.total_amount),
            order.get_status_display(),
            order.created_at.strftime('%d/%m/%Y %H:%M'),
            order.shipping_address.replace('\n', ' '),
            order.payment_method
        ])
    
    return response

# ============ EXPORT EXCEL (optionnel) ============
@login_required
@user_passes_test(is_admin)
def export_orders_excel(request):
    """Export des commandes en Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, '❌ openpyxl n\'est pas installé. Installez-le avec: pip install openpyxl')
        return redirect('dashboard:orders')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="commandes_lozashop.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # En-têtes
    headers = ['#', 'Numéro', 'Client', 'Email', 'Total', 'Statut', 'Date', 'Paiement']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    orders = Order.objects.all().order_by('-created_at')
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=order.order_number)
        ws.cell(row=row, column=3, value=order.user.username)
        ws.cell(row=row, column=4, value=order.user.email or '')
        ws.cell(row=row, column=5, value=float(order.total_amount))
        ws.cell(row=row, column=6, value=order.get_status_display())
        ws.cell(row=row, column=7, value=order.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=8, value=order.payment_method)
    
    # Ajuster les colonnes
    for column in ws.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(response)
    return response